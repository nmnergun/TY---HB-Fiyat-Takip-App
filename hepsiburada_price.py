#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Hepsiburada en uygun satıcı/fiyat bilgisini yeni bir Excel dosyasına yazar.

Kurulum:
    pip install playwright openpyxl
    playwright install chrome

Örnek:
    python3 hepsiburada_price.py
    python3 hepsiburada_price.py -o "/istenen/konum/rapor.xlsx"

Statik Excel şablonu kodun içine gömülüdür; harici bir kaynak Excel gerekmez.
Yeni dosyada yalnızca başlıkları
"En iyi HB Satıcı" ve "En iyi HB Satıcı Fiyat" olan sütunlar güncellenir.
"Atanan Satıcı" dahil diğer bütün sütunlar statik kalır.
"""

import argparse
import atexit
import base64
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
import io
import json
import math
import os
import re
import shutil
import sys
import tempfile
import zipfile
import zlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote_plus
from xml.etree import ElementTree
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

BASE_SEARCH_URL = "https://www.hepsiburada.com/ara?q={query}"

# ÇIKTI AYARLARI — raporun kaydedileceği konumu ve dosya adını buradan değiştirin.
OUTPUT_DIRECTORY = Path.home() / "Downloads"
OUTPUT_FILENAME = "Paralel Fiyat Takip Listesi - HB Yenilendi.xlsx"
ADD_TIMESTAMP_TO_OUTPUT = True  # True: her çalıştırmada ayrı bir Excel oluşturur.

# Statik çalışma kitabının sıkıştırılmış içeriği. Harici Excel dosyası gerekmez.
EMBEDDED_WORKBOOK_BASE64 = """eNrtvAN7ZV2zLhx2bNu20WHHts2Obdu23UHH6ti2k45t28m3+tl7f+/7Puc5v+DsynWtrJV11agxR91Vd9UcGVNeChQMFegLECQQ
EBARkOFFtpkuMBBQHBgQEAoQJIiWkK2Nk4mNk56yu52Jow69m7UVYQEYCGUeEAjQ/8r/01KRNCbbx4go8ij+JrRHJ7lWJwDJdh4w
hEytwNfdnzj400grx9z8Myf8FfFbPpoFRy3y1Jon8/WEWZ2N9a+AsME2RdhI/+FuFV0cDoaBoTOyLgysBAdNjcs8VBwBJiw9vWjT
TKA0+6Y+RVbhbUHJCmwzNHQ+v93slSFVpM3BFtifmqVcw6s9W7KqG0ico2VziLszKPMW5lSeure7dBJgkcSea+guuj/kidSVqAwV
WxxeUqU6HSmbWRb0jUA2GEPiOuDXJze/tHVy4MuwYteVRdz91ByMs6Tp8tL28uiijIm8k39pZUcz06PTyhHwJwqY/Q0kbxnZgF7J
bDPl3c+PVig9dtVLuP5Kto2C27+JY9yaExtnLItJbE279WMoR1iOi4M6wKdkeKSZs1vu3AjXy6Z4BmIPTOTpE/UB6bj1tvsb4kg3
Eal36PnW8PFXz7Ppfa3Ai+uUY1/mjy9AQJ+foEDy/xGWjSqMJHeAd9KAqIMBhKWeg4mVIwP9n9f/Dcf/lb/CMUFGro8RNuRS/C10
+HFhPQ7xm5SroHgbkbKPn8QaQmN0ISnOehcFLqIydiComG+Hz9vo6jpP9Mkw0cMOa0UkMupXkl+ulSGpL/bOENTkDimi9hXb/Ugu
C/cLD+liku6ULh2x1OXNpflSI5skonPXwrWIU/Rq2FAoP3BD0HnT/TUHZTQ2DJ+qgBe+KZ1XhtQ0ceSdJE5l+i6lFB/D8YdxSZ/N
pqMNOxp+X67yGrYiun/GVlEy4zKvMcxx4CZ21zO26r4VzuxCefeik80aViRmRX5inhIbIsu799VsYL0ey3bv54DqKx++GgmIJCjd
AP6nIPneen5QDwoE5AwFBIQA+IubFYOrrYOloa2t5R/eqlC1ti3kRPzwmX2H6HfFwJEWl59Dq26VPP8pWX++XjIKriQPfBhlrqjy
uukq/W02dY5qdZ2ytgcfn5/ldjxujiGzmD2n+Yx6ZlLGIUAmsCl9uPzBwMZHq9jMAjvb0HrSYuam5Xn86+3Ycq5vVlWwVzvGzgyQ
36/MzB+KNJt1rRV7cKey/RFNNUqQy4zgTTc9J6oJ9e00XGoy1suYchUhKpUtIlf7TJgFlD+jhohPcTZ+0oXtsSkJdNGUq3AUsC0z
2cElxr5DmxRoAPd7q6iduAOxRZPfThFoEi17w9nYhllakXpRXJFgfvHkgfrFh7ugnI6aTTkPJaClNUvwRvGbeE8Gkp3AyA4XuZHU
96cT/GPF3l5iBblnxQhw1JaeKAlhpSANH9hlotlFTqZ1Tja6yoTtr1m23xb3yWzcSo3m8xVKErpUOVMyCDYDGV0ace/UZ+XYLlAP
w+MNkw8OnQmikzCLSxZQ3/C3D4ujrSGMrKF6IS0GTMZQY4XLBUK/dvmoOIK6KEPLN9dufT4k78mr5zFcoYmmPJ91thn+ZPXQmJI5
1YKY1LLeVK9Vb2r026nTLUvwpvXPnSYba6hHd2szaejwueRCTPwo7aAhg2vmi2Ug/Fg6g6oD5ppyHfCo6jtfa0y/G1mHgQuYlznH
kuzlFAiahQwOP83XP2lJmxx2Cjl2tROVfZOu0U1uE7Y8apEgKif4Rb9sukb0+7L8TWAX/JJPfOoyxP6yzUW9qSEzorotkPSusral
ckW7AiI2Z+h0BOxNCwGnSaVqtOVYP/Ni6WNnQQjKKlzsnnhYtz4jk0HJM0FYN46Q4w6yF9h5mrWfSZLkGzJH+rIsKoJVhZpTyCBU
Q8iyGUStXv1hANLRdGsZPDA73PEpy567bAhYTKxJxaTAmgc27HewVSXrHo6TAqU+2Naffvj5nFdXUcRZnPQFc+nRxfO1CsJGOA5J
/RJHJ2tHg5efMQ8cw6sEm5QZJiNFmpIpvL2eZj2Rcc6ExZVC3Y9sy9PB01wC8cWkIknjjNB9UCI2WdxVYjuU9nOgGuc+1yMHLI0h
JVxKvKjFVSPbG0tGrVrrDOirwTo5G/PZXWHOa65yn7nbbG56tyTRc5kkcIXVqWNjJNYMCfvr1h4V9nZQA17t3jXqYJ+0iV1DJBWf
sKX39NtM+SxnpMAOWWHEJz/wM4p/i8Dfm+3NMvrn+Ii7dUzm8KXaZmoWEhcufSK4RIsf6gW+tvHLiQ4t2MBsoAsdw/Wzved3Pph9
5wf6p/j250tKvQW8awVwHTYQJDAgvv+LB/89yv+HE4EBnAj8/zwnKMpI9TMidsB8hDK6ZrqgIS07MRCfqZwiQI40UK2SE7H43G3i
ULfqqC9KMYDahrhfTgy5PLLB74Ci3PfQlEYgg3mgWF1+BbeduJ0AhTQ8r8QPEANHQT47OS0/h4HPcERp4UsYps1B5Tx4oiTDCi0s
Jha+a5kNYbWYZHRmtnCcxhsFcMJYw57fS5DSWWVAQ1MTYQaV6f5ni9Dqw5fcAHeCmVssKAkeOHegmaN5IpqnHpEHyJv2N1MZ90nL
nsIX2H3ccin8n2y8XiFOhKHDYgS+VAmO+wXpbSqX207fXkepulZvaMfuLk33RoYWNv8RMr+f06Gr8IGA6FuBgDD/RQmOZiYmTo4M
f/1i+oOaLB1L2zXuhPfO7g+IvtfnX5lJmut4pFjCtc504G3qZ7f1u07Iuyv2iw/eBrWf3lSNysbZSuPWk/qmCayxkj+4d3ZrxE/u
Nz/cmvlfz6/uvB6fzidatCqnpuA3bQj4171fbw7e+D8LX81dfR7eHG7P5ybed5v6Mm/5EaYOfT1fYz+vXh6HSvRyuvDhGb739DYd
chzq6el2drU2u3W+e7g1u+vy+vB7wZDlEuTi2+j28ud2avP4+nKveWXQezR3XZ06O17bOT6ev5ltL8xRsePz+3a28WrzbXy82D09
OLDeODV3uUN98Pt2vH60q3s28w6s9U3MMXT5fB4/XG97Ml5fX3xc7E+NHD0v952Xf4w1DZVkxj5ut3za3U+cLxPIbXfxvVVr2rC8
y7p6NLd5e3j5vL3VQacwxPDd7l4yPGBwNBmG536SslR1unG/uXAgY8o8d02RNSFp8WGXf8gaRWPi9N3fnSy1ydlOfj7qqDirLXzc
lBw46H3uab9flDzczL7PO11mTLS6X7aZ3voe2ur5tsLAxso5Pzg9vR28ZXve97ScPw+sLs3h3Om5hm7mJnbldqx5vT32vW8/7CNe
7ExODWLc7e7uMpDUIPBXBk4cQjGutLd4twBc8WHn+3nwJmR3+exb3BbUFYKjPYw328Xgelqs9qKXO8m3wUOwwb/x+bheklq9nZJ/
4sXt1fl8/0H2zMTJ/PlkG9ulR3DR6eUNNU3eb+SVr+v9cntz5YWzsbe1ORr72OndjsAB/en7YMexOHxFvPFm75W3692u1VKsy+Hj
6PNwffU2ux6NPjVl4+vt4+6lnpUugxc9Xx9/dXhZXz9Brf3xdnP4ZBMaKceKWpXeDuVLH7JmofbsvA+9BiXHrwH4rPp2NImQo/9V
D/kjKpU12ABSZtlp8+rQ0Ljtp2SH67eRp4SJ7QrZmN6B+K1juLZ29xb315fj68d6+cfAxsNLDiOpE93oR0a9uCvTqN+uJE8T7qWF
/q+HV4afDy7vLztNV782JsNSkGRJNlh6Lfv3kZAsQtd+zIbR/FKwFB4KixZ98xIOclW6mDXrZyP5hT3D0mu95Q7xYFxS6cVoeLCL
ogdWEFSJ+DK3jN1jO9mHQy/XDPELvX+4aJVD/o2PbmMpM1moOMbP1AOFXd5toMk17Bh9/DdItO3vqUER/OYOVV2Or25cHg1c6jHd
IgSKmdfZqrUEmaeJlqOQi7Wu6irow5heepeMT+a1jhN8ys6lFGL9uN4VxCGgOBWkMKIFttp7PdWW+0s+CAO8K7i0YowvZRPr5Cif
8ipysZ+QIU9CA1YYw6ifcB1v5tcbS+V4wpc3kJwhXf69NLsfOGb+reAk2xz0N78pt0V4PxjVmwb9ShBSoR5/K+S+tcbqQ3KwQpyD
YPXhcfN7rc8Zqi2sLqBnjIyRGPqs51+N2umD5p+iR/Wv3mBnPJ3OM3StAof+/3OQvusj3L87itUP41ur0yFMPyoZsLnkoPFajvUz
urI0G6mKerUljFkPiN/CwVi4mtQusXmWmWWUzhaJZYcYQzqzVspM3wrd3jjuIehB9ILUbRMY3LI9oet8Ma+1ffNpHPJ5CBu+qwcs
QrNIkz7py6DI/PzWf19BKq93q3S7WK2vgxJLANGGPNUqOPNAW34xOV3VYptHIwRoqlj/2y+h7eRtAVC2Ai1GZNvdAmj6Vtn3n7wO
vwsnehzAuFsQ/lzAI9kMwLi3mSlgADY2nAgfzbMT6wG4d6R48yM6iLxWp7+WX1mwqCp59f6LWLKZI6PVzh8Xph/FvnQI+EWbhfE3
NJwb2/dLg52Vtijs29TrVVOb0dN/H5YGTtIN96ek1Qa9PgEjJVVPVcgUSr2/G9PHKpguLxkf6JR+w3fW/FmC60R/XppH4TP4ewle
bOhSpWWzmde5gYV4bm703vxI4pVetn8ap8rTbKkq2VZFSbDj4ocX09e0DX2YovHWoki1nVE0qlUouAWgxkr94uv3SYPq8tHATgdp
EWpb82mj5+ktRD2BeGihaFV596rGy3t0sWxjF8YOImpZ+ZYq8QNZM5mgEioiMic44qQC1nDxZHoUA0pp8oCGqu1WklqavGOygpc1
KfcKlBQOYeGX4ZbBMvFLblVvfPBq/eKYfKKB8tEOPRw24oV26aB8HhxUA4dEdZKoMXEos012LbcRD+avmEj2cEj2AYguIpJu34Si
Baho04wQom3XPVUu3n3GZJSoQ9IhhYLYKYf3yOMhT5mPtgGTdJbpAKMyJCH+gkRhtsPkq8CobMiUNv5TFqQum3YwHxklH6xUoh8h
uPa3XJvlvClIhLB6AP5CwM26Uv1K4nLPXYgFYKYSLD89T4mDcWe/c5oNcK2rMrZQ5wvXjQhTJ7RIqism/wQbMe5zFmuVeb545Exc
lXI/I4W1bbaWJOBK512URjX/PpSAFOpfqe8gXZQwa5OV5DU4QtwHETrhTpGnTiGrML3iWthGDmft3iddChX8tGhGHuHUYPQcplmK
Jh0Trqy7Il2kmHEaRGTQNI3vtCNs3NscBSE72mShU2lxJOFRx4MpyOQRgSLcyIC8gILdaoW6EDBilz8qQ28qLPobi6+ZbtISUry5
hp6ELd97dVc8j6I+0CTFIpg5E6v/hTiYN98fUvl3AJfshcKKdLJaIwwDftGEqXAMpBC1rb2DkA1yeA0qe/Fkg+3XLJfjkTN94SDL
IOZMkgJax3FiNEu3UyUmjQXWhfxRHZ0FmqyZ9iyidHNypjjYnF+cS0R5w0FkNVTBFVl6oWcWcHE5tAgpQqyXRfhfxUTjBx8qKO0h
Z6FNM1x3uPhVLnTWK35NaH4Imb8r2xphn49eDW3DilDkDUcSRdFkBc5ObV9t8Kbj0dqmJaHc/WKYXUXR5FDhLFyDNiUdCJYFVu0t
vRkMhg6gU6FuYnp1Nmne+kXZBI7qXrPHbsqbbNnG9tO6jlt9khGubMWFfgTpRIReJMBTG/iL0s648E96OvJZh8RwZVsWlBA7U3J5
zE8rN3xk4hpICYFdKWNGQYfVulHpUr+9dErFtRwTGpb4UalIKDL+Su54aBvTXvnJcbCDaF92Ho40cR2WNIHyQJPkdHngERaFombe
0zFoinwwlkZZIjT/SiVq/iYD3Mahe7H64HEQ0j6IJpEBurzfPBrp5tji5ubtyxXsbTWo5s8IFbOFmbAptuPjUOmWeqRdFKxO2M3y
QoxDDeA/5FTc1SnrtYjqU1lIm2inEGebF4sT24R1ad+jIxMbHNkqU4jqA5GQh6RLx92IFZybFEi7ect/Qy9ZQ5IPiVCLsGgXQWel
imtC4x5CDEQOB9svwx9CNN62Z/mNw4uKw9LLxm2qxL/lTMvT2ImE6jNLlzpkhmMllKGFoNiEOhKBIJNXjFGf9TY6rvLGFA+f6af/
0MwfVx25Ee6waBT3V6M1YkIAeCQDelkR5sfougvGIkZedxoRF32F8NfJ4GEdYoxBE7RihiTHDSX88D8Gm1WP0kiVzekjhN1MIDVC
I63BWGVClTl/XOkmx3rCm3EbixoXC9FSULZQ5w0Hk/UqnFFtWnQ50xC2UPQFHiMcMCZtk06UjVAibycT9drHyDMqkQzQRKi4x/FP
WERH2IaMo4Gq5WDJcmqjbUBRJhhQ2WpblGWvzq6h/gFSc7ziyLswiCV9RIPrADT6qTtNWlm8Q3urrsSMhdej3+TYcv84eJLuH3Vi
gLqDQalt1H2jAx9sBo4m0krifo9D2LF8bIeN4W3pM6XiVnJ+tz1bjH5xay15Y/5onyzqhiUziYN3CFGP/dPyiAwx3qCJdCjyWD4S
XFuTgGHcfg9kxIx1Vtx9VQTrU9BiINX4a0PAiH28pGImDFmv/W7sduDqdn+SHzFGiEuJiMWoq9puaRMgcvfhliLv4t+5ucvLReIh
q9VHFptJxRi5qBXhWmHaiB+aaQx/xWnQWT5tuVD2gqOW1QhzTubTbOpE73bnjMOncEgHzz6MFGPJiMRfVam370AXqjIObeAkol2q
uLut1VCGetgmPxVvrMpPf82rB8CKPmqsRlLP8pDrWsYUgoQ0S0PRlGUk6qe4i5yZWa4dtDAjF3nS2YNlVy/Y+r4VMQubCacgSZCs
QDyUkEt9t/wkWwNm0B+r2jTKfLxLAgsivZ7fQDE4wYnTYDF+rqVrj8fBYoSgwKM5UBBu97ZGoROZHbo0FAYAVDR2vqVPzEe0/rGk
D1IiAMYSbdbHTpRWrbLPF4UtcTwRFjXL4mXnmQWMJE+6k/Rjf6Up4kbC2KWmPR8NkpuUL0ZIWU0Bo+wCCk457wh9XJmK57xGfn87
LICwtbY8jYgF47cKpmyQjnvNgI5APIyQ827crubXpupHGunZO2P3GlxDbll1Xd2/BjDOO0GXjolS9r/oYwZbOWfWGqYVYvSyS4M+
BkNtqRHmmBSl2tQJXB0G0JP6hh3oD2HXOg6NpUjiGZ16uwE0nUp9cureHGirAqcP10v3w3UrYjUG2Pv2ITKiHZG9XuR1Tr2qLFe4
qsFTQG4NCgqVH7Vi1igso5xdZYtcxL7/pe+gbCbzteo+9lRwcfpMb3LRgTeUqMBeyoQCXkQ1aoxnAXH/MCQGhxAQ9HiByfNWdRmi
9YpF7ZYp1Wk5NuCJ6UTK+zZKt/BRWPFGvXHTnsobDonDQfZYD+pLK028knEAUIgKw5LPGwvZsKqQtluqVrtRS9Gk4aJoBmtK4yJx
kcOCd4ht08kKk8uO+Fhums8pygGAQPwFxaNKcMVGhyLCyXViBc0x7y9uG9Frk32z62jZDuzTSPofFIh7w1CMCdKsAbj0JgRCkXoT
DyI0I5x4XafeZcBA6I8pyZWMSqWM8+L7gxAEVqUZKEpmg2GbFYJiLwRDfMeKPz41AKBM1pQEDS2TeJk3/z4PXe7AHJwlnFc3q3C0
z44V97c/tFGIsahiCMr2O3gwEQBMzeMzXi0Ok6FleqOVfxxKi4L3UpuZtbOKGn4JBeeV9wRwaazTpmtILAnpEMTYUQ0wynNe9HrI
5TH6qXnqaqPDhmcGS+pN8EEP85hx0RTuL95lvMFicW85sy2QhSZ9ZiEtFXwXy3c3s6+FLOvN23Jpu9800ddxysOTP0R78t/30r9R
xTZgmqanKAeXZC44JaoU9XHQml775aoycmkmUb5Lac6SHc0nYn4E/YATH3PoE4fhD9Ez0hTiFcMk2+4E+qOsmGvOZQD1UiweL2dW
StOsJcrIpeiE/VZ9ItmKny1hkHxu7y8EAIEirL0sszyvGCG46CyAIKOOruLUMmVUBU9aFYsmOWgt7Xda0MbnszFD86kW2qyjSCJM
UVougrLF6zOUuqbDMeIjlWa3R92bNoMSocJjBmQrSEJMqkiOjTxTYUvYCERavfp20A+XNXv2nuQjG6z5ZpM44QerYCnesPfz/IgJ
Sor6k2+JXOt08lfOR270OY2VY+9qy42/FhQXimfZqm/veE5DmxYYyUS6KresRt+e1jiaiu55OD/0VDOh0HCEW6mnbs86nbetBnPJ
KgcbuzMGGNbEn9xSuw+w4oiAePdRN+UH4iixYU9K9Wno1DLVmdRfC/nbPCGe8LkcRKLEw9EwyGa3oB9tW/HySIpd4GdS9zFyqEgW
DOWHHhSIF9iqH+6UaQKQySYk4VRv+iZfgMPnJsZs3CKxt4MaXgWAFwpZa5IYOWz9OKDEGwkRHBBWKpqmaMDX2AaurtQXJbcQg1lR
bqnuAxHiVrx1S9Eqjmb2unL7AGQrg95m2goVf16TlSY7xOmIGHuisUCT9GH5lMT2AvJC8RyOiih7YAeaNKNPmXAKFXz4VvZh+RNc
EBeTpy28EzhgEeaoNbe6vZvSyWRbn1gWLYDfAPVik7wpy9yJTDx7LoB+5TNhyMWCkEvIkcSEq1RrQODb0vOrESmIkUqO7K+yVSPj
lnvCwgT7oTSui8Wj5cwIQep3AkaQ8uebAbUsWvN2oOt0kwuglsVUTGeW8d32BERtH05gGAqqm1JRNwUroj+0NuNQutBKXv6oEMyU
0TJWfDlSyMmgDnH2GimyQakNnw9XNt6VFp+vO1Uq/irAkrAF+xIrmvlFyALZAzBgTYXNrs2R7IX3whhJOSaVggC1SuYCCdKiSRr4
UuecBs4fJIZhChfFZpEUiFtx0KrbR6oAYh3Jgu6XBknSYvFkhvcVJ89pf7h4UDJYCehIsm07VNZfuYhYsZ6LOMLYpU8TLdbFApwF
qZbMoGS0D0sHZ6OaVejgh1Jz1IQdMS5NeMf9jrlxr2TArqpRbYf24hxDoL3vRmB+DjPGK7HCeo1ff2EWALaG1XUbLmdKKtEEqc3M
WVduxGYUeZieNOcytWJMG4n5Pc1dSK93VI61DMgzcFKi2rgb0OQYnmQF579LqDOMpmSfSNXcs7zRGAzP6NCiWbBXnkeDGSx6dEmZ
/xTdzeK53H0Dv8RF5cx4eXpvMzV7jJvl911wWIWYreO2g4YcBb4CQwFKnu150umZled7l66CxAxAmGgUo2WNOjj4hc/OrBLjAQpw
OGlHGFNjo/ptBpUiqk1qK4I9E4j6cGndXrveXY28bD9inMVFmwqs0edWKBb6ZuNWirHAbUTXqVGHKgtU82M6GEWe4mK08XhYbn2Z
ZLXdVTFLAVPxGFgh6d2QNtkL5JEa+sKf8T83oku47AAtFBFTGqegtG1k/11QiIW/oSaAwxVNjl23ObG1Yz25MtPS/ajuaPrC7diB
YZPAUNMZ51l9B8DntnixsOdBe4HOA31e9iAJ1jHRqA7pxuR4AgGG59kwcycUWk7PW3W3IR8d2Jlc4RhjvU04kVSH9OcVdj+stA84
T9w0GFLD1X/B2svZBahvoIfvQ8HZ8tATYEjTP1AL4sUga29vc6XWh3M9TKvrYclOA7tn++3g2BbFcz1Q1nkys2I+HhLFR3yrWmeY
HdJCT2/GLUcbx3wNKTr1VNFtmZ7a1h7ahx1pZA8RGEQh+9oiFdVBFnddgXO+lpPAA+hQvxk2vJYJyC4qnJKXihPZapjs+PMDQKxc
RKVB0Lnc0HuuCjBNgpbG0ALASDNeOYp80Q5qsPRfDfbaZukXC4Mn7iZpUy3SBIoGm2CXNHd0afo36jbLUH60EVwf1TZ8BJqT/khM
H5p7aTZBkzBe+AqPv1Z6WHyCfeTBSActQvg2JZgCn2YO+k8ZTsCiQoQ2MR5MlHGdFS3bHqatXpDeLJ5d7jI+oIj/atpH+zhhUUD4
3lpJbKRWlEHBqgy7lkyXRoLVgKGerLYa8QgyFTau/Ws7cAi+EyXOzkB7MHd+pvMawnjkLK+qe8a8WB5FnjTJlZb9mjRdlTGReOs2
RcW98foLML80JyAF8tC2AL7vnh5h5SkY83gxkf9QGp7T8BuvMVMjDaVgbb9N0wRc3l4Za4PQR998HcBM/SU38EBcboKmaueS5UfD
rqHwCoEzOFJe3VH6u2AbMRZPvj9wUk8AFhZkmaQ2ZPGoq8o0WHOsk6jxF6EQBHwEpTMwVKwaZcYHoNGmmlNIChlAqePW11HLnFkO
LTLqNjSKiRBuDp5IQrwQxQ5IdzGQJCVRENmATpe0d5jK8VSoNw0WThQ2ajqXMMpVpBfQ+OMJGNCmUd15xaRTcyBYZSDV04cch9D0
QcRafgFMdBqnNW32eeurTdblH0NGiPbf0N3VigCuIwfhuGsEpMl4atHMMNbIdiJi2PgomGNue2Jke0l0jddOonjpv8obJArjeFcR
bPdvoEhuQemAbFd0MsMjY/OLcQmSR2byqqdw9AIMNZwR3dIxGlDfPOMdmXkC/L2BKVyavNJiEtdG3AFYEuj5sQCsMkia2rJA2cbd
KugyoTVWbq3hvVlYAq/oMcI/1bnDaHO/yOukWCigOroazNsJgr2tEqf9k87lGFGkY8KUOVeE3xh7rJ62rkHZAXkAKUWVnDipdF+o
fvi/r8zDXY806U9C/6bBqLTbkI+dP3ou9G4qTh5qyrnNBLcEJZ7GSL+LC6LzJ0GahKmalJeMRyk+/rk2oTOmdNo07TMlrOPF1biu
nj667yjptQmAxIXr/VeSTI9qFOC6xNZumv4zvRBL4LSJc7tVjLLDP0WxB6CCgiWjmpgDaLVE6BsHXxuxZ7shjTDuseC9/Mw+BJUb
BqBR4LkC2lGM8DJy7zboC7Q8H4DU/DUTrJXyFo13Hb72QcV+d3IGiUDNShAp/gNFxoupdFy0YG0t5aK0R9qv116pqoxiA9I4wsrP
ejlYjIUsiSUh/PhLoE4gEayYJ1gQaksrch9JYOl482jKKLLT06/EcLz5+cR5Iejjwl8IcktHHzPPRxW64qr6yY9Nn7bszz0o8sT9
krWKfQTJOl3v3iH/8tp7LxS2h15RPgAgvfnTVuBJRsQqSQb8Lk+QqD6jRuzxSLYhGPiuFK4gEcpZBQXJtzZiQxkA/o2AInNOWtYG
VSkqeBT9hT7lYg7OqjBzJa4ioULQCSWgp5Ius1fKCMaEP7gAqLOgaRaiMnMabxviCgFtHBpKMbH3AS6aCLWHrbgCz9QPNPmYCGUl
7swI1nPpXG/mdsjQ/t5eEmARatjD43q5FfdQPDkAh9OQSwd7HIw/P70SE/Hm1wZ/I6PLw8WQrFzRpqSKnpMCTDN29wTQljAWsWaD
V2VxwJOR/LF1sANGhbfyUjN39DrRdGUu/TgKPnIbWm5ro/dft+U9nx5uDrw6bg5t9eQiZzbxtsW//DwO464T/CrKSgj74caxCjIC
m0MrxiOr5IxEFpQmJ7AM0VlzmPvlHZPTbDT5Z1K9FsbxMQnubmwEd0Hxk5roW3JxTudtJN6DIemcjHqO2xu8LaF+i8I3v2SvwZxv
/N+e+rHxY3pSQPFCTH6irTgUN9mdEpt4B5HQlVaSWBT9WJnGl4ODvCGmXRiMNadP4yKDGFBLocCf2oJu0Cd+/n6LJG4QGPp65W9m
BSYkkk/baL6mT56t/ZtCyip/FFXCpondiSanbE5UOQWHrI1lragVB2Efcq2GAl9FYusLhiGEjWz8hpdBHoh1sKThb+bS6UTYEVnM
rQLxgLZGIt1r4exuEWqOCY8EEId74R9VLTmYLZa40hD0jDB4zYz4/h5O75AGeMAcemZcjEbq7GkSuR2yN9kC0vG4rpZC8ztMjf5p
yj6q2wKIvOje1cDb+qNXRMTSeHRp4295++FDQVqEX5ZsIAwr1tgfaU6QxC9DqQNTQJlCTDJ4lqZyztUi+lu0cEAwpnigGyif16dT
2dn02+4roDDEfzVxzK5hBnSPBkh6eO5Cycfr9qFcuYb/dtAfqLFSxsDLfOITcKZSPCGCESRIdHX7UNgSMbePTEWN296HHLkxLHXm
Sae5qstD1o2iYHSbCbftKlS2K2U045E3wuQolYguGZoAYRMZXa+Eu4+EIAa3UMzGfpC1kyoOstrvlYedjNGyMkaD/rpjL41YqiEP
xTaIMp0A9g4Wg7CuP+aP0Vc3E5cdB4OQ6GXSOjPd4wYvrU+mUkFOIzm4WlHEnANPEN3ECTP7wAW5WiOMhHAOXUH7KKHOlAP0PVJI
qBh3bRZa9heuOhMtTkVWwYFOekbuFqHWVfvPRSgLaNVoJZFLNsIwicKqcXYHZxEuhzT15xLfFjcIJZ1r0Xj0UyS7r91sbsNj3Ms3
iL4RttAlZQ372tX8k+Wo0nXSdkXB5BpgyT9Ewr0yUwiq6BF0NGcn3v3Q0Z8XSaa/IxW8XyNVKNSlaUAiNGAS51pGQsumpDDHX0bh
sk5KBa0ULKbemlNhI/IdoN6oMirZtegf6BJEVBeENyD1KTUSU5fRqWLjPRrfBvN+ooYxWtlTVPfFG6c4UVyOsr6DhWIHUAyCUSsi
aWkoUyZn5hMv3HH2oJUysnoH5s6GlwZBS80laivEIv4aSoxBFVAUklV8GnX4FZI04vmdUl169ItmUnuTmcKND0vOKswUnoU49WVt
JlwD5k1nT/YNUvK0IQ+IHELkM8A3QurTWtuMFuIvDZnG4sBsRdLTX2mkomEE+4b2D7/4ZNzvdNxiI7a6DEk8gCdqeKPJFEWj2eTG
Aiawi2RXInO0MicGCUPZHFKT1xRghOj4rW69hVmQv6fWcghNQKyQhFoqUrFxge/sXNSZedM8hyJkT7iUMCUv9LczUnKNIZTHxTdF
iL4U1jWxvhShUIeEPIx5zEyrNhXyw+zk9kHwM1n53WwOVRMJcOt91/ZM29p4PAgcB/ylGDBV9Ukcua3rtEh5taRAz5XN0RRQmmO2
pN3RJPzMqsjD6tC+sf2JsYnVpU3bTtyP54PjsfWmw/rXvkOqakxa/nXv59a39/3r1zRrXDmlp6enm5unmrfT0+2DuW3zN7WjZT5+
vk7aT0jb/YHLWLku6Uf3189/3OOWoCLe64IAAkIh/vMvu3/tcTuZmVib/NfrX9vbZ5oTsThNaF0YR58C4zqL40eeBlzTJRwPFx7b
yTp74hJJIbqTdgKKbTSowNCHSVJ+QBIzUv4bHpQBRWZ+XzruZj/7/aLZ30lehrHMQ2ioH15AZSLXO8blux6fMh/xCTZ5JtdjqwxD
6MPikih0GJp3PcNIWfhKRupOedOjaKwRJIITuUx5iHRsZYoeZzr0Pnbo+eSRwqFAnoXxt6owqJLxlcsdvvSpb2xTSWw1gWMwqdrR
pCy0dKfmcZguTD9mLmDO9QtBxQieFw4ETEJmmM34vi23NSnwGzKKFRThpzHZFyejSA3qHx76iw434jL60zOeejDOZenJDiffIPYp
9iDoTHH46t5vrFPzDQenrZBoJeBkdmnTY6TlhprK1mbZUNyLGbThW93jACm1NStofdreYwHlTGxJsNCbZAsWHjf14WqBmKTCr2er
ytdfk7034k0KliwjnKb4gI7eLbnm2q9hweGgb/gxexlBbN72+ZSYbsNR9nG4Xt0rVKPF0LYDi0SRH7YdmZav2xwc0Rn3LOejbJiz
vy8yhmJQhoVYOx1TofMovBrSDU41Y4IKgGKpsgpaZCeA5wjiPioOkGEDWt/2B/kIFHqD3wiBQN+qL0JXfY3svQTsJDRvQcgdvovf
CjqQm/1s3QCdPkkmZPB92h64vdTj+TyNnOLzvH7TchJOjR9V1nAI5dRI2fd52039/Hi3ne76uAkl6HpZiU41eSu35fO9b+r6TPJQ
twim0v14aFPb9H28cdV7u5XT42p4BMtJ2UbBRgmXHSbzGDdyEzCCzuRPj8iUkCaWpoPC9bcJPYmPoN0awWbBqJKzQuFDizPR7n4t
/46+VC9egeo8/sYYQbfFdFgxFTwHhHCvTQYjpzoyy9uinmwfy9RpN8HbHb6/NUMdfCepQw2BVGuWw5xF2U4KpSD+OTgFZc+zTpGk
6cJbzNJDpkd5ByXiLZ1Nnj5dhiYePE0U3cNgiSFo12iSuVtSOpV+Wha61yvaW9c9LujzLc1k8NXth92ZIlcPEw3OG2T5sMiNFZ2p
yZo49dVdAmOMqCPxD6mDeL9DOq53opD+A6dXJq0hdkVKj4tZ9yAuBpLQNy/EbU5nh2IJ5UEf2ikiYSbu/VFw/oMH6JnHLYisM+dH
JUla1y4+JB5/MLViZbCwDxMy3M52ueYO/CUr/3CasTthK5OlnLzJo6PqFFUzAqTeK7IclIAcrzP96cMyk9uAu0DIulEie/BFUKuh
n51yYfPByPNsbVQ3YZI0xlPDCZfbju2dKW6LnOG1v0uewdptgi+Uuw2d5N36TQ9NtqCCQRB3BsyTxJCdkgI8G+ROzYMqO6Q2pVzb
q4L3uzg0jGroc9eddAJgW2phSR84fk2AsVgGEOuTpxWQqLWCI429B4dtqEX/HqVqqsGw3v5CGVXhQVE8chqRi0IL1vV2Oc9HUHFi
T+OT6vFdT5erXXX/rO8dYZFdqF1SFsqKoi389FdDC511++frrOmmZiedmQ1MxmrA4D30gyB0RkGI/02x2X3aoFx702V9XlR2GRuJ
2GyDD2n74vgQo5mHYjFTyzloMZL2yaE/FO2aoj5JIJhaVwE1d6Qd9Ht+txNTKrNaRxLlPvKKqOlvJ+URtvXi54aI2GceS9eMj4Tc
dHlDlahXODMek93pn4pYrcujEGxdcM59LuaalTggmg/kpJKJy6QW8MiVE1uW4keTy68GM/Y1Zpb3OGrnaSCl1h0gl+uQG4ttuoU5
TAExuPNHjucl4ws4lxaH7WpylyzkNGJADeiuvOsz2o98OT6ttPMb3XPoULtJtBVWyINqxd2JSO1whzMaF22fyt3r80vZvT/Qea0j
JiwozRpWCHqXuMO1x+Boqb5Dp+RJqe8S0XIjTNit3KNPc4SMcQZvxb+sw9ST4XaNJw2rvq0WpSiLs/Uzj9La1Mke5KZREyMW2i/N
TXXBQaSkkSbEn3bPvrckwNwHvLbEg1av6pJHgqeX1W1KkazMLtMVPdFqcPqy7dgiLCZ4gV/wRR6kmo2q7jCFIFo2gJyvcyXbeSEa
V+UocSKsmtYRf1SdS7s/NMG3RHhFaWBve0lRfb+FtWFPosNzbGoUPKb3NWU8qoIqkero67z99frSoalV9Zvg60Cx+Q+0DuvW+uPQ
1X2pigidecPk/YtTrqZ6CRfE0knrnfXHD+RGQ+bIDNP2zFDn2Ukxa4vLWmLKc74JXpdVG2RZqfbzRQSHpULVS8TM3lmwTi5uhlzr
22MWcXwhThHt2rOGTDK775Jjd8qYm7LPXtl6zOlkGeOHrZQTis9GdQJOPgv339T20wjwDxeApICsLXTVDbQEvzd+hzmZoMeYQD9T
5aQxql4o/3Wlaug43orGXimCYe3cyKWw1NiarvGV1tMiMST6x5bFUsQ0OQam5OEZI8dYR9z3KmVd8z40TaiVe59zDo97wd+Yym8u
RsplSyc+P9aPFqg7RAvRo7D9+c9Cc0F40QYESV4fDGBLBK2yfYENBl07A8I4IHRdTXmqHG6ThRbcNr3AkYImqwnvxODxOZHNggvz
8B96BU0YYTOSPTitVFcqNauZITP7aiAyXRzXXueFML7g3oX2XTcjqWaAdi+86xbOni59STb1nEcY6inluwPOsuVXu0IvdLSxgWlm
bb94h/gnvk9JL7GLAQUCkoEBAoL7L753dHK3MnH8w/Ttqlq2q+yoPgQ1nwJXOjPrHOr0e5wg0E4DfsbduubhCeGKRQIzZWv5n96h
ScY0lnVFF6odYU2311/Vcu7bqg6O4n+mSMTrfw72hUVoo0srQ33sEjmiYzZmhAXYUZkFyDUMXLJ3fdTZ3IfPvYL+wAYOMkPxmm7V
aLHJJDCTWcFZlgf++kOZbYZRJXmVPSKhgO0bmyvoVGYLmwHII2Q5ty+MuMx8idHuXp1NmKnnYDqfBm3Uidd9J7JiPC/Ez2yCeFzn
D2nwUqVdMpf7TE+LFJrlUcVNoCqasYT9O61WKdfjhLcAsOEloufaU3A5p8v5ZPH48iMSlou+TGeqYBsF+thvFJwgBYNg5XLffNDT
l0EQEErYTx9gPEhxVUHObBZb9hSJ4KpAx55NReIyHihmRcfQvUklR6ZMS0jZdkt8qjH4EBe+Ed65sbSZ9ukDlWJxDoi3hujLOX7r
1n0UAtPtskJo4i+p/fkANck85cOvdh6+7ksOcxqIntXNyT8kadYLgs5uKU4tkzysHxBy3Vfmn4cY99mBcN6iJqsWkfVG6NMmQMNF
Axt5shiyo58AyzfFwvhagoMR24IySt5Py/sKSveoiCa/PpjFuL3JYL2YlYJth1YOf1BtYmQTrykE/rsCRpSaCpJD4askLKjfBNXh
YzHIh0ZOLWj3hSRpjLPq6aNIQH5KhhtBaLcl5oPbO9iVHvKN4IcIkatkPgXrNnNXBMzJqJoE1o8cCucU0SLzJ6r9CmQ2tuRh9ewo
Jae4CG/gtrSdxD6+cykhtU6r8K3K1Axqjvnq1IqXdfXPcu5uU2y4z5rtUlm5nr09mtHgz/z8TpxTD6Fl/ZCrbNLFxAbuIW7masUA
/ozigL5X9n6aFQKsJemswkd3YLDFCxHCRu5Gbubh2S5DX4xic7fTcaFN1wLWrK/H7NFZ7uyxBzM2OZNvA2paw5OlAg1H9Om2p4ns
OqGVUN9P4mjA6UbiT0NvTQMYdjfNemLY4rBl3yfxf/xChOqM4k8D2Z+CBV05RqDjybsSExvwe1t36VxbN/Gvv/oFQwFtfka7ez4Z
Xrh6bpNK19Hd9s597O2Td1tbib1XSlCpCf8ej52EN7ypt4jMk353V/nu95D0ypzmbn7mKYvuof+EpcEZqYH5VeMMok2joSOy3q2X
t7NO6EZ789KTTohP11xVabL/qHxGaE7FDln6UwTK9S7I8ziQwxf0nwKXU8syHBcQ0q34QECo/x24ZgYOJsZKTg7mNt//it8gjTEH
QPyuuxZ/ds9ItxlJ8zxJWn8njExR6kCWiMe5VE+v5UqhVCvisfch9J1WKdYxPbVqPZE98b0PSLIKxvDD0MUMffR3JGj2g9fzzpqq
aDNOL4k8KxXU5YSjGSmOHGNt47Zox8ldEewbjeRsi/oeSctlsUbhvrtwHcuPe+Bu4G+VQsBWFbnzsN8jyNche2D/YWifaeznpU57
aA/8SV8Nf7Nqgqx3sCShOP4dYt6+PdA9yxTOslXdZEAh0NOmT9PkmKB8r0gzn88al+Dg+NrAfk1Lm7y+L3Dn6UqAXjnSVrTk/scw
domS1Dj75sfrapBGZuibaTwazGnJy3eQX9jfFynp2TeLM22dm+tbdYwX3VT6pcAk2XUPmq5R1QPOD0Zu6M1xxdKTN0oWn94LZWFa
Qvi5ZC53Cz7WzV9msGkrP0CIApp7WmokYBRAxN1sHp3DpgV0yy60f1QHpblV96y4lKJYNlnPKoeJsPtHcfJbrlRwMW3RZUKibDTq
4F1Ivncd7zdxzo14xjgWx/tYKh2IX3jaKjBbM1GLX1ZG1NBNBiE95Ppneox1LtJAfkFLO0MH63HUL6c+6jdV/gJsVOIiPukfyD6q
AXVnVRx3DD3lOq/RFsj5qyHQC28e3aW0e9DWaiforfKcIzAfO5cT3v/OupOlW0fcgOWCPzDzDrStYeTuBfjwyz1QZoVSC1egi5pF
iccpmY3fqk6hO4bd3IH3b68iNQRGjr7ijsAG2wT2oXWI36R0HLKwVVJrAhWfsEyF39sTm3s6EC0WYTPdIMxRLg20ZKzk0HGhMJ1E
Rc7583+S3wW9cGuM7lVFlao6UpUvjGfVDSENSCi9NA6UIT0Ga2xgiLoDdyzf9eBQojGFT5t3OSLacY42cM5lFWnZ+3kvCKr38KXO
Sre9Q3P0bn21qZVbkRnMh6AKBaJaJ8XV53VEwz+JGntAWFF5V7Go1jeJXe6Z3oraaf3STGE7X31c/dW+3jl9YfSiC43BXgTBM+D9
e/+C2JKsLnc1t9evZtgKfw8/mXk3U3O/MkGubtH91kDr09bvxJ0q5CrCFNwWYwhmt1g8zFvy1dKlMbkbLTJcsI76Ogq2X+X9gYUD
WJaFatZa8MxCt/GKsXdpsinY1VzYMU9BDrV521Sy/a+q8I5wIPx/GZvhIi+yw/eSWh1JBvUhWaKGnqk+o0dzIFDcGEK6Hg3IeHPi
32kFTrZFvtxIaK5pdbi5J+ZF6jqyaZlc83T2q80dWGBtdpOhbOJMfQm88DRoLJg5demyOfNO+EGd/6FPXBGPaER61bqKgkFm3Ljk
q+u0pMd9b9v9bO6Vim6PZ1aZQvMjtgDnvnI4xSncICgKPtEtpW+xCj3XlKafRPvwcvZ+yaRAeV3X97hpUY869Xr9K/2seXu/cmyq
hIt7l8AlfLYb1neC2fRZVYqc9aQpCHvgWOruKf5c01Pf+L6OUoaKpV2WmYE3USF9ELcjEWe9sWhYdt/GDSrQeve0dXWMOFwNnbhP
3ClbX/PFVJPfJVes8Wve4vGRyNq+Mttle48mXtmT8YxFgy+3B8B8mM4FPjTYi13NMbeZC7MJBqAiLFYF1DExzYjv94etTvzYucud
u09Houk+0vTJem08xpKE+kk0wBeIYlDyDWtFFIjSb7AqLxJCP5X7eAKQq8btG8q2wfAFa1QIXevc1X3xRe1iHxAxTDLkbeq7Ky3X
EoXV9+tjvWiWNBDnnZMlftCvhh/FDlTbNapebCoxsBSz4/3kJROeoM+v0ozF8XB2/WIramjmyCAwzKdAtOIhAeU/JK7Sn4azmCrC
I+XoJitzy11nKcTevpBvteYY0AY/uqMkjnBhNnxXtmIpyXaHMjnv3NtMHoMrpj9ltGPBQ8iqvIR3vxDPHhqKVBDOQV/v/WincX4S
cjTKLxyCLyWDBx5A9pPa0X7BBjhG8XomJDLqjLBK/Jl1E5/rPslz/pBIKRlBWkljTQhKEQ0RTgFB8hZko7W+3Dil5GjPnXRFiQJk
PqDOHlE7fFMENXn9FwIV9nsnboSFGKzIp9MtEEp7vn/BTqi182BBIU+IlCmvzNS+MEq4H7FhVzkjxPOQCAG+pYRfPK+g6AJan3dM
2OV6dIRTVU6+MSgxX9XBpbfyEGf/Mu2Cu/byQUAcSesjx55RwUqgjeetPsvSghOOc8RzqXus/omFGLtQ/u8ckHzyin4i/wNGZ3u3
XyigOVbUNLC5AZVmAUPllyfAjs/q7vAb380pCavszeyupBW4SpcVthiVtsIKJ70Ogk5ifpKtfqd0zjxl0Hn+/qrN8o1xIkg0rVix
7IsdFR/Lg9wlUK7CK0rIomDMM1VkjF2ShgMevKvM4ELie9lA/kmpBzfU1XVHrnZ+5E4rmhPIxO+S7RjzQeCdAtVpnWEi6rY67vLX
BS/opY0OnFHgtwevt38+DeSmUGOTAwwERA/gWiQgSGBjWyN5B1s7RwYjWweT/z6n/r+HgAASpThms8qI2A5xxA/ZPjZPYg/bq+rX
7mEeUGA/Q8iRssZp1Q8x7Iz6LvVJzq8qWWfydSSf42Ly4uzM7vB7dausbjeEyUTAXmIUqX5Q8mgF+t7hDXd2PCkhR8LSYndAA+Kw
xv2+GT0uymKfqy2emzJiwrAxCIqNYA4iXiFrGMU0j7RhcPhW4X2GEex7BDd9zjO0WbyJ0k/glr1hc7G01vh4K8ofmFTrK0u8/qio
FtP6jNNeCXM6dAFL23Mh+Mvl89BaGK4rw+uHKYZAQydLeYO9xebkMstjYS+OHr6351V6bCrFt5JAZQeKjMrqQmwlbd+9gu8v+Qnt
xGMFwYIgYhbVmiwnZBzud09XN1c7XWGcUviFVtUZhKz4wwy7ggjbHTD5HQj5xftPMiDaw/lfSwVymi4PvPrb7PF1lV2x+pp1SEUu
9/OwjJHXDQJTtfbU3O1b2slV4tHmWL/KRBqmfz8zHo2JJWithH/52hQStHH0KjeIoppSkXXOZAP180OAq3f0xSWOYbC0Syfw2u0i
btdjt7gmt+iBof3lH3u7ylum9HoAYtUBoEb8d1Ab2Nn9L6b/JVmJUn8wHfII7gER8JrmoAkWLMhklo2uMV8K/HVZrixeJIkYPAGp
Qjdv6AjLIaT4cVT8G2rLyM829+tc25Cidp7FlEbMgmv6wzGRQSzuc5hdeK6uZV+u124N6S3omg2Dya9Rb4od3iRNPRyYCT3D4r0D
WWpUbDREBVU1Ic15DYXSx0Jw4Wt2iJ200YM01ZcWThJZ35eDu42+qenbL0oHeClrNx77isBUe3a6tV1ig6jeGM2pFxLNJFkYOV3N
pIZbC+CkelX8PiwUpRlol0Y+hL/XYSxTSkd1VLtgVddWglEJiKYccBTQJbk7XQ2AUhhhr9ScH4FAn5DT3OL6vklkiJByix4AnApZ
w4CT3fNLRlrw/Rn8Yfvj8hL7Nzz3Cdj2ti70T7zbQC5+9qYJz0Zc6GFOujNwUXLJrNkLmTrMrmQtyGO0p31cH33MGbEsf/szrkOc
4hLZTzPb7h5gUDJyy/PTL9pi+SMvFmuxbnNn72ikXpoP5vIX1JDq0+/SqUzge2znXOrIz8JSl0O89xQcVTR2ELQ5pLrkLxxOX31Q
G9vVjS+ZPVDr01Cn9L7xgvwP0IFB6ID+b08a+bv8w3NH/j7Afz4T4d8lFRA9//aEhL8r/v2c+L+k/cv/cWr878r/eQj1PzAJ/Zfy
/+VI6t/H+fvJxH/JFOz//Zzi30f5+97Pv6Se5h93gv4+wN9vJv1LCln+dmvp76p/b2f/Jdoc/9zc/n2E/6Twf5effEBA/weh/139
P5Plv4uv4L+p/3fqlJcC//LnO2jAjx/AeWbCfz79f3jBdcM="""

SLOW_MO = 0
PAGE_TIMEOUT_MS = 35000
READY_TIMEOUT_MS = 5000
OTHER_MERCHANTS_TIMEOUT_MS = 2500
DEFAULT_WORKERS = 4
MAX_WORKERS = 6
MAX_ZERO_RATING_PRODUCT_PAGES = 6
DEFAULT_SHEET = "Sayfa1"
MODEL_HEADER = "Model"
SELLER_HEADER = "En iyi HB Satıcı"
PRICE_HEADER = "En iyi HB Satıcı Fiyat"
STATIC_SELLER_HEADER = "Atanan Satıcı"
SELLER_HIGHLIGHT_COLOR = "FFFFFF00"
HIGHLIGHT_SELLERS = (
    "YILMAZ SAAT GÖZLÜK",
    "YeniSaat",
    "ONDOKUZ",
    "NOVA SAAT",
    "Naim Usta Saat",
    "MENA ITHALAT",
    "HOOLYA DESIGN",
    "ElizoraTR",
    "Duru Saat",
    "DAKİKA SAAT",
    "ALEN E-TİCARET",
)


@dataclass
class SellerRow:
    product_name: str
    seller_name: str
    price_text: str
    price_value: float
    seller_id: Optional[str]
    seller_link: Optional[str]
    is_main_seller: bool
    is_cheapest: bool = False


@dataclass
class ProductSummary:
    product_name: str
    cheapest_seller: Optional[str]
    cheapest_price: Optional[str]
    seller_count: int
    product_url: Optional[str]
    rating_count: int = 0
    comment_count: int = 0
    verified_no_match: bool = False


@dataclass
class ScrapeOutcome:
    product_name: str
    seller_name: Optional[str]
    price_text: Optional[str]
    price_value: Optional[float]
    summary: ProductSummary
    status: str
    error_message: Optional[str] = None


def normalize_text(text: str) -> str:
    text = (text or "").strip().lower()
    text = text.replace("İ", "i").replace("I", "i").replace("ı", "i")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_model_code(text: str) -> str:
    text = normalize_text(text)
    text = text.replace("_", "-").replace("–", "-").replace("—", "-").replace("/", "-")
    text = re.sub(r"[^a-z0-9\- ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def product_code_tokens(product_code: str) -> List[str]:
    code = normalize_model_code(product_code)
    return [t for t in re.split(r"[-\s]+", code) if t]


def title_matches_product_strict(title: str, target_product: str) -> bool:
    n_title = normalize_model_code(title)
    n_target = normalize_model_code(target_product)

    if not n_title or not n_target:
        return False

    if n_target in n_title:
        return True

    tokens = product_code_tokens(target_product)
    if not tokens:
        return False

    if not all(token in n_title for token in tokens):
        return False

    compact_title = n_title.replace(" ", "").replace("-", "")
    compact_target = n_target.replace(" ", "").replace("-", "")
    return compact_target in compact_title


def parse_price_to_float(price_text: Optional[str]) -> float:
    if not price_text:
        return float("inf")

    txt = price_text.strip()
    txt = txt.replace("TL", "").replace("₺", "")
    txt = txt.replace(".", "").replace(",", ".")
    txt = re.sub(r"[^\d.]", "", txt)

    try:
        return float(txt)
    except ValueError:
        return float("inf")


def safe_inner_text(locator) -> str:
    try:
        return (locator.inner_text(timeout=1500) or "").strip()
    except Exception:
        return ""


def safe_get_attribute(locator, attr: str) -> str:
    try:
        return locator.get_attribute(attr) or ""
    except Exception:
        return ""


def safe_count(value: Any) -> int:
    """Hepsiburada sayaçlarını sayı, metin veya boş değerlerden güvenle okur."""
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, (int, float)):
        return max(0, int(value))
    digits = re.sub(r"[^0-9]", "", str(value))
    return int(digits) if digits else 0


def deduplicate_sellers(rows: List[SellerRow]) -> List[SellerRow]:
    best_by_name: Dict[str, SellerRow] = {}

    for row in rows:
        key = normalize_text(row.seller_name)
        if not key:
            continue

        if key not in best_by_name:
            best_by_name[key] = row
            continue

        existing = best_by_name[key]

        if existing.price_value == float("inf") and row.price_value != float("inf"):
            best_by_name[key] = row
            continue

        if row.price_value < existing.price_value:
            row.is_main_seller = row.is_main_seller or existing.is_main_seller
            best_by_name[key] = row
        elif row.is_main_seller:
            existing.is_main_seller = True

    final_rows = [r for r in best_by_name.values() if r.price_text]
    final_rows.sort(key=lambda r: r.price_value)

    if final_rows:
        cheapest_price = final_rows[0].price_value
        for row in final_rows:
            if row.price_value == cheapest_price:
                row.is_cheapest = True

    return final_rows


def extract_hb_search_candidates(
    page,
    product_name: str,
) -> Tuple[List[Dict[str, Any]], int]:
    """Arama kartlarından tam model eşleşmelerini ve değerlendirme sayılarını alır."""
    articles = page.locator("main article")
    if articles.count() == 0:
        return [], 0

    raw_candidates = articles.evaluate_all(
        """
        (cards) => cards.slice(0, 40).map((card) => {
          const links = Array.from(card.querySelectorAll('a[href]'));
          const productLink = links.find((link) => {
            const href = link.getAttribute('href') || '';
            return href.includes('-pm-') || href.includes('-p-');
          });
          const heading = card.querySelector('h2');
          const favoriteLabel = card
            .querySelector('[aria-label^="Listene ekle:"]')
            ?.getAttribute('aria-label') || '';
          const headingLink = heading?.querySelector('a[href]');
          const title = (
            favoriteLabel.replace(/^Listene ekle:\\s*/i, '') ||
            headingLink?.getAttribute('title') ||
            headingLink?.getAttribute('aria-label') ||
            headingLink?.innerText ||
            heading?.innerText ||
            productLink?.getAttribute('title') ||
            productLink?.getAttribute('aria-label') ||
            ''
          ).trim().replace(/\\s+/g, ' ');
          const text = (card.innerText || '').replace(/\\s+/g, ' ');
          const accessibleText = Array.from(card.querySelectorAll('[aria-label]'))
            .map((element) => element.getAttribute('aria-label') || '').join(' ');
          const ratingMatch = `${text} ${accessibleText}`
            .match(/([0-9][0-9.]*)\\s*değerlendirme/i);
          return {
            title,
            href: productLink ? new URL(productLink.href, location.origin).href : '',
            ratingCount: ratingMatch
              ? Number.parseInt(ratingMatch[1].replace(/\\./g, ''), 10)
              : 0,
          };
        })
        """
    )

    best_by_url: Dict[str, Dict[str, Any]] = {}
    for candidate in raw_candidates:
        title = str(candidate.get("title") or "").strip()
        url = str(candidate.get("href") or "").strip()
        if not url or not title_matches_product_strict(title, product_name):
            continue
        rating_count = safe_count(candidate.get("ratingCount"))
        previous = best_by_url.get(url)
        if previous is None or rating_count > safe_count(previous.get("ratingCount")):
            best_by_url[url] = {
                "title": title,
                "url": url,
                "ratingCount": rating_count,
            }

    candidates = sorted(
        best_by_url.values(),
        key=lambda candidate: safe_count(candidate.get("ratingCount")),
        reverse=True,
    )
    return candidates, len(raw_candidates)


def hb_main_product_data(page) -> Dict[str, Any]:
    """Hepsiburada'nın sayfaya koyduğu doğrulanmış ürün/satıcı verisini okur."""
    redux = page.locator("#reduxStore")
    if redux.count() != 1:
        return {}
    try:
        data = json.loads(redux.text_content(timeout=READY_TIMEOUT_MS) or "{}")
    except (json.JSONDecodeError, PlaywrightTimeoutError):
        return {}
    return data.get("productState") or {}


def hb_rating_count(product_state: Dict[str, Any], fallback: int) -> int:
    graph = (product_state.get("productStructuredData") or {}).get("@graph") or []
    for item in graph:
        if not isinstance(item, dict) or item.get("@type") != "Product":
            continue
        rating = item.get("aggregateRating") or {}
        count = safe_count(rating.get("ratingCount"))
        if count:
            return count
    return fallback


def extract_hb_sellers_from_page(
    page,
    product_name: str,
    product_state: Dict[str, Any],
) -> List[SellerRow]:
    """Ana satıcıyı ve seçilen ilandaki diğer tüm görünen satıcıları çıkarır."""
    rows: List[SellerRow] = []
    product = product_state.get("product") or {}
    main_name = str(product.get("merchantName") or "").strip()
    main_price = None
    for price in product.get("prices") or []:
        if not isinstance(price, dict):
            continue
        value = price.get("value")
        formatted = price.get("formattedPrice")
        if value is not None and math.isfinite(parse_price_to_float(str(value))):
            main_price = (float(value), f"{formatted or value} TL")
            break

    if main_name and main_price:
        merchant_id = str(product.get("merchantId") or "").strip() or None
        rows.append(
            SellerRow(
                product_name=product_name,
                seller_name=main_name,
                price_text=main_price[1],
                price_value=main_price[0],
                seller_id=merchant_id,
                seller_link=(
                    f"https://www.hepsiburada.com/magaza/{merchant_id}"
                    if merchant_id
                    else None
                ),
                is_main_seller=True,
            )
        )

    # Bölüm geç yüklenebildiği için kısa ve hedefli bir bekleme yapılır.
    other_merchants = page.locator('[data-test-id="other-merchants"]')
    if other_merchants.count() == 0:
        try:
            page.wait_for_selector(
                '[data-test-id="other-merchants"]',
                timeout=OTHER_MERCHANTS_TIMEOUT_MS,
            )
        except PlaywrightTimeoutError:
            pass

    show_all = page.get_by_role("button", name="Tümünü gör", exact=True)
    if show_all.count() == 1:
        try:
            show_all.click(timeout=2000)
            page.wait_for_timeout(250)
        except PlaywrightTimeoutError:
            pass

    other_rows = page.locator('a[data-test-id="merchant-name"]').evaluate_all(
        """
        (links) => links.slice(0, 100).map((link) => {
          let row = link.parentElement;
          for (let depth = 0; depth < 7 && row; depth += 1) {
            const names = row.querySelectorAll('a[data-test-id="merchant-name"]');
            const price = row.querySelector('[data-test-id="price-current-price"]');
            if (names.length === 1 && price) {
              return {
                sellerName: (link.textContent || '').trim().replace(/\\s+/g, ' '),
                sellerLink: new URL(link.href, location.origin).href,
                priceText: (price.textContent || '').trim().replace(/\\s+/g, ' '),
              };
            }
            row = row.parentElement;
          }
          return null;
        }).filter(Boolean)
        """
    )
    for item in other_rows:
        seller_name = str(item.get("sellerName") or "").strip()
        price_text = str(item.get("priceText") or "").strip()
        price_value = parse_price_to_float(price_text)
        if not seller_name or not math.isfinite(price_value):
            continue
        rows.append(
            SellerRow(
                product_name=product_name,
                seller_name=seller_name,
                price_text=price_text,
                price_value=price_value,
                seller_id=None,
                seller_link=str(item.get("sellerLink") or "").strip() or None,
                is_main_seller=False,
            )
        )

    # JSON gecikirse ana satıcı için DOM geri dönüşü.
    if not rows:
        seller = page.locator('[data-test-id="buyBox-seller"] a[href*="/magaza/"]')
        price = page.locator('[data-test-id="default-price"]')
        if seller.count() == 1 and price.count() == 1:
            seller_name = safe_inner_text(seller)
            price_text = safe_inner_text(price)
            price_value = parse_price_to_float(price_text)
            if seller_name and math.isfinite(price_value):
                rows.append(
                    SellerRow(
                        product_name=product_name,
                        seller_name=seller_name,
                        price_text=price_text,
                        price_value=price_value,
                        seller_id=None,
                        seller_link=safe_get_attribute(seller, "href") or None,
                        is_main_seller=True,
                    )
                )
    return rows


def scrape_product(page, product_name: str) -> Tuple[List[SellerRow], ProductSummary]:
    search_url = BASE_SEARCH_URL.format(query=quote_plus(product_name))
    candidates: List[Dict[str, Any]] = []
    raw_card_count = 0
    for attempt in range(3):
        try:
            page.goto(search_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
            try:
                page.wait_for_selector("main article", timeout=READY_TIMEOUT_MS)
            except PlaywrightTimeoutError:
                pass
            candidates, raw_card_count = extract_hb_search_candidates(page, product_name)
            # Kartlar yüklendiyse arama tamamlanmıştır. Tam eşleşme yoksa aynı
            # ilgisiz sonuç sayfasını iki kez daha açmanın doğruluğa katkısı yoktur.
            if candidates or raw_card_count:
                break
        except PlaywrightTimeoutError:
            if attempt == 2:
                raise
        page.wait_for_timeout(500 * (attempt + 1))

    if not candidates:
        return [], ProductSummary(
            product_name=product_name,
            cheapest_seller=None,
            cheapest_price=None,
            seller_count=0,
            product_url=None,
            verified_no_match=raw_card_count > 0,
        )

    selected_rows: List[SellerRow] = []
    selected_url: Optional[str] = None
    selected_rating_count = -1
    top_search_rating = safe_count(candidates[0].get("ratingCount"))
    if top_search_rating > 0:
        # Arama kartındaki değerlendirme sayısı seçimi belirlemek için yeterlidir.
        # En yüksek kart başarısız olursa yalnızca sonraki iki aday geri dönüş olur.
        candidates_to_visit = candidates[:3]
    else:
        # Kartlarda sayaç görünmüyorsa ilan sayfasındaki sayaçlarla sınırlı kontrol.
        candidates_to_visit = candidates[:MAX_ZERO_RATING_PRODUCT_PAGES]

    for candidate in candidates_to_visit:
        product_url = str(candidate["url"])
        try:
            page.goto(product_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
            try:
                page.wait_for_selector("#reduxStore", timeout=READY_TIMEOUT_MS)
            except PlaywrightTimeoutError:
                pass
            product_state = hb_main_product_data(page)
            page_rating_count = hb_rating_count(
                product_state,
                safe_count(candidate.get("ratingCount")),
            )
            page_rows = extract_hb_sellers_from_page(page, product_name, product_state)
            if page_rows and page_rating_count > selected_rating_count:
                selected_rows = page_rows
                selected_url = product_url
                selected_rating_count = page_rating_count
                # Arama kartındaki en yüksek değerlendirme sayısı sayfada da
                # doğrulandıysa daha düşük sıralı ilanları açmaya gerek yoktur.
                if top_search_rating > 0 and page_rating_count >= top_search_rating:
                    break
        except PlaywrightTimeoutError:
            continue

    rows = deduplicate_sellers(selected_rows)
    summary = ProductSummary(
        product_name=product_name,
        cheapest_seller=rows[0].seller_name if rows else None,
        cheapest_price=rows[0].price_text if rows else None,
        seller_count=len(rows),
        product_url=selected_url,
        rating_count=max(0, selected_rating_count),
        comment_count=max(0, selected_rating_count),
    )
    return rows, summary


def render_console_table(rows: List[SellerRow]) -> str:
    if not rows:
        return "Eşleşen satıcı bulunamadı."

    lines = []
    lines.append("| # | Satıcı | Fiyat | Satıcı Link |")
    lines.append("|---|--------|-------|-------------|")

    for idx, row in enumerate(rows, start=1):
        seller_label = row.seller_name
        if row.is_main_seller:
            seller_label += " ★"
        if row.is_cheapest:
            seller_label += " | EN UCUZ"

        lines.append(
            f"| {idx} | {seller_label} | {row.price_text} | {row.seller_link or ''} |"
        )

    return "\n".join(lines)


def normalized_header(value: Any) -> str:
    return normalize_text(str(value or ""))


def find_headers(sheet) -> Tuple[int, Dict[str, int]]:
    wanted = {
        normalized_header(MODEL_HEADER): MODEL_HEADER,
        normalized_header(SELLER_HEADER): SELLER_HEADER,
        normalized_header(PRICE_HEADER): PRICE_HEADER,
    }

    for row_number in range(1, min(sheet.max_row, 20) + 1):
        found: Dict[str, int] = {}
        for column_number in range(1, sheet.max_column + 1):
            key = normalized_header(sheet.cell(row_number, column_number).value)
            if key in wanted:
                found[wanted[key]] = column_number
        if len(found) == len(wanted):
            return row_number, found

    raise ValueError(
        "Excel'de gerekli başlıklar bulunamadı: "
        f"{MODEL_HEADER!r}, {SELLER_HEADER!r}, {PRICE_HEADER!r}"
    )


def create_embedded_template_file() -> Path:
    """Kod içindeki statik ürün listesini HB başlıklarıyla geçici dosyaya açar."""
    encoded = "".join(EMBEDDED_WORKBOOK_BASE64.split())
    try:
        workbook_bytes = zlib.decompress(base64.b64decode(encoded, validate=True))
    except (ValueError, zlib.error) as exc:
        raise ValueError("Kod içindeki Excel şablonu okunamadı.") from exc

    if not workbook_bytes.startswith(b"PK"):
        raise ValueError("Kod içindeki Excel şablonu geçerli bir .xlsx dosyası değil.")

    converted_workbook = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(workbook_bytes), "r") as source_zip:
        with zipfile.ZipFile(converted_workbook, "w") as target_zip:
            for info in source_zip.infolist():
                data = source_zip.read(info.filename)
                if info.filename == "xl/sharedStrings.xml":
                    shared_strings = data.decode("utf-8")
                    shared_strings = shared_strings.replace(
                        "En iyi TY Satıcı Fiyat",
                        PRICE_HEADER,
                    ).replace(
                        "En iyi TY Satıcı",
                        SELLER_HEADER,
                    )
                    data = shared_strings.encode("utf-8")
                target_zip.writestr(info, data)
    workbook_bytes = converted_workbook.getvalue()

    file_descriptor, temp_name = tempfile.mkstemp(
        prefix="hepsiburada-fiyat-sablonu-",
        suffix=".xlsx",
    )
    try:
        with os.fdopen(file_descriptor, "wb") as template_file:
            template_file.write(workbook_bytes)
    except Exception:
        if os.path.exists(temp_name):
            os.unlink(temp_name)
        raise
    return Path(temp_name)


def configured_output_path() -> Path:
    """Kodun başındaki ayarlardan varsayılan çıktı yolunu oluşturur."""
    filename = Path(OUTPUT_FILENAME)
    if filename.suffix.lower() != ".xlsx":
        filename = filename.with_suffix(".xlsx")
    if ADD_TIMESTAMP_TO_OUTPUT:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename.with_name(f"{filename.stem} - {timestamp}{filename.suffix}")
    return (OUTPUT_DIRECTORY.expanduser() / filename).resolve()


def prepare_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
) -> Tuple[Any, Any, Dict[str, int], Dict[str, List[int]], Dict[str, Any]]:
    if not input_path.exists():
        raise FileNotFoundError(f"Excel dosyası bulunamadı: {input_path}")
    if input_path.resolve() == output_path.resolve():
        raise ValueError("Kaynak dosya korunacağı için çıktı yolu kaynakla aynı olamaz.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_path, output_path)

    workbook = load_workbook(input_path, read_only=True, data_only=False, keep_links=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"{sheet_name!r} sayfası bulunamadı. Mevcut sayfalar: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    header_row, columns = find_headers(sheet)
    model_rows: Dict[str, List[int]] = {}
    updates: Dict[str, Any] = {}
    seller_column = get_column_letter(columns[SELLER_HEADER])
    price_column = get_column_letter(columns[PRICE_HEADER])

    for row_number in range(header_row + 1, sheet.max_row + 1):
        raw_model = sheet.cell(row_number, columns[MODEL_HEADER]).value
        model = str(raw_model or "").strip()
        if not model:
            continue
        model_rows.setdefault(model, []).append(row_number)
        updates[f"{seller_column}{row_number}"] = None
        updates[f"{price_column}{row_number}"] = None

    assert_only_target_columns(updates, columns)
    save_target_updates(input_path, output_path, sheet_name, updates, seller_column)
    return workbook, sheet, columns, model_rows, updates


def sheet_part_name(xlsx_path: Path, sheet_name: str) -> str:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(xlsx_path, "r") as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relation_id = None
        for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
            if sheet.get("name") == sheet_name:
                relation_id = sheet.get(f"{{{rel_ns}}}id")
                break
        if not relation_id:
            raise ValueError(f"{sheet_name!r} sayfasının paket ilişkisi bulunamadı.")

        rels_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        for relation in rels_root.findall(f"{{{pkg_rel_ns}}}Relationship"):
            if relation.get("Id") == relation_id:
                target = str(relation.get("Target") or "").lstrip("/")
                return target if target.startswith("xl/") else f"xl/{target}"

    raise ValueError(f"{sheet_name!r} sayfasının XML parçası bulunamadı.")


def column_number(cell_reference: str) -> int:
    letters = re.match(r"[A-Z]+", cell_reference).group(0)
    number = 0
    for letter in letters:
        number = number * 26 + (ord(letter) - ord("A") + 1)
    return number


def patch_sheet_xml(xml_text: str, updates: Dict[str, Any]) -> str:
    updates_by_row: Dict[int, Dict[str, Any]] = {}
    for reference, value in updates.items():
        row_number = int(re.search(r"\d+", reference).group(0))
        updates_by_row.setdefault(row_number, {})[reference] = value

    row_pattern = re.compile(r"<row\b(?P<attrs>[^>]*)>(?P<body>.*?)</row>", re.DOTALL)

    def patch_row(match: re.Match) -> str:
        attrs = match.group("attrs")
        row_match = re.search(r'\br="(\d+)"', attrs)
        if not row_match:
            return match.group(0)
        row_number = int(row_match.group(1))
        row_updates = updates_by_row.get(row_number)
        if not row_updates:
            return match.group(0)

        body = match.group("body")
        styles: Dict[str, Optional[str]] = {}
        for reference in row_updates:
            # Kendiliğinden kapanan boş hücreyi (<c .../>) önce ayrı eşleştir.
            # Aksi halde genel desen bir sonraki </c> etiketine kadar ilerleyip
            # sağdaki statik hücreleri de yanlışlıkla kaldırabilir.
            self_closing_pattern = re.compile(
                rf'<c\b(?P<attrs>[^>]*\br="{re.escape(reference)}"[^>/]*)\s*/>',
                re.DOTALL,
            )
            regular_pattern = re.compile(
                rf'<c\b(?P<attrs>[^>]*\br="{re.escape(reference)}"[^>]*)>'
                rf'(?P<content>.*?)</c>',
                re.DOTALL,
            )
            existing = self_closing_pattern.search(body) or regular_pattern.search(body)
            style_match = re.search(r'\bs="([^"]+)"', existing.group("attrs")) if existing else None
            styles[reference] = style_match.group(1) if style_match else None
            body = self_closing_pattern.sub("", body)
            body = regular_pattern.sub("", body)

        new_cells = []
        for reference, value in sorted(row_updates.items(), key=lambda item: column_number(item[0])):
            if value is None:
                continue
            style = f' s="{styles[reference]}"' if styles.get(reference) else ""
            if isinstance(value, str):
                value_xml = escape(value)
                new_cells.append(
                    f'<c r="{reference}"{style} t="inlineStr"><is><t>{value_xml}</t></is></c>'
                )
            elif isinstance(value, (int, float)) and math.isfinite(float(value)):
                new_cells.append(f'<c r="{reference}"{style}><v>{float(value):.15g}</v></c>')
            else:
                raise ValueError(f"Desteklenmeyen hücre değeri: {reference}={value!r}")

        if new_cells:
            first_new_column = min(column_number(ref) for ref, value in row_updates.items() if value is not None)
            insertion = "".join(new_cells)
            cell_reference_pattern = re.compile(r'<c\b[^>]*\br="([A-Z]+)\d+"')
            later_cell = cell_reference_pattern.search(body)
            insert_at = len(body)
            while later_cell:
                if column_number(later_cell.group(1)) > first_new_column:
                    insert_at = later_cell.start()
                    break
                later_cell = cell_reference_pattern.search(body, later_cell.end())
            body = body[:insert_at] + insertion + body[insert_at:]

        return f"<row{attrs}>{body}</row>"

    return row_pattern.sub(patch_row, xml_text)


def ensure_yellow_dxf(styles_xml: str) -> Tuple[str, int]:
    """Sarı dolgu için bir diferansiyel stil döndürür; varsa tekrar eklemez."""
    yellow_dxf = (
        '<dxf><fill><patternFill patternType="solid">'
        f'<fgColor rgb="{SELLER_HIGHLIGHT_COLOR}"/>'
        f'<bgColor rgb="{SELLER_HIGHLIGHT_COLOR}"/>'
        '</patternFill></fill></dxf>'
    )
    normal_pattern = re.compile(
        r"<dxfs\b(?P<attrs>[^>]*)>(?P<body>.*?)</dxfs>",
        re.DOTALL,
    )
    normal_match = normal_pattern.search(styles_xml)
    if normal_match:
        body = normal_match.group("body")
        dxf_blocks = re.findall(r"<dxf\b[^>]*>.*?</dxf>", body, flags=re.DOTALL)
        for index, block in enumerate(dxf_blocks):
            if re.search(
                rf'<fgColor\b[^>]*\brgb="{SELLER_HIGHLIGHT_COLOR}"',
                block,
                flags=re.IGNORECASE,
            ):
                return styles_xml, index

        dxf_id = len(dxf_blocks)
        attrs = normal_match.group("attrs")
        if re.search(r'\bcount="\d+"', attrs):
            attrs = re.sub(r'\bcount="\d+"', f'count="{dxf_id + 1}"', attrs)
        else:
            attrs += f' count="{dxf_id + 1}"'
        replacement = f"<dxfs{attrs}>{body}{yellow_dxf}</dxfs>"
        return (
            styles_xml[:normal_match.start()]
            + replacement
            + styles_xml[normal_match.end():],
            dxf_id,
        )

    self_closing_pattern = re.compile(r"<dxfs\b(?P<attrs>[^>/]*)\s*/>", re.DOTALL)
    self_closing_match = self_closing_pattern.search(styles_xml)
    if self_closing_match:
        attrs = self_closing_match.group("attrs")
        if re.search(r'\bcount="\d+"', attrs):
            attrs = re.sub(r'\bcount="\d+"', 'count="1"', attrs)
        else:
            attrs += ' count="1"'
        replacement = f"<dxfs{attrs}>{yellow_dxf}</dxfs>"
        return (
            styles_xml[:self_closing_match.start()]
            + replacement
            + styles_xml[self_closing_match.end():],
            0,
        )

    close_index = styles_xml.rfind("</styleSheet>")
    if close_index < 0:
        raise ValueError("Excel stil dosyasında styleSheet kapanış etiketi bulunamadı.")
    block = f'<dxfs count="1">{yellow_dxf}</dxfs>'
    return styles_xml[:close_index] + block + styles_xml[close_index:], 0


def patch_seller_highlight_rule(
    xml_text: str,
    updates: Dict[str, Any],
    seller_column: str,
    dxf_id: int,
) -> str:
    """Belirlenen satıcıları En iyi HB Satıcı sütununda sarı gösterir."""
    seller_rows = sorted(
        int(re.search(r"\d+", reference).group(0))
        for reference in updates
        if re.match(r"[A-Z]+", reference).group(0) == seller_column
    )
    if not seller_rows:
        return xml_text

    # Önce bu kodun daha önce eklediği kuralı kaldır; çıktı yeniden girdi olarak
    # kullanıldığında aynı koşullu biçimlendirme çoğalmasın.
    cf_pattern = re.compile(
        r"<conditionalFormatting\b[^>]*>.*?</conditionalFormatting>",
        re.DOTALL,
    )

    def remove_our_rule(match: re.Match) -> str:
        block = match.group(0)
        if HIGHLIGHT_SELLERS[0] in block and HIGHLIGHT_SELLERS[-1] in block:
            return ""
        return block

    xml_text = cf_pattern.sub(remove_our_rule, xml_text)

    existing_priorities = [
        int(value)
        for value in re.findall(r'<cfRule\b[^>]*\bpriority="(\d+)"', xml_text)
    ]
    priority = max(existing_priorities, default=0) + 1
    start_row = seller_rows[0]
    end_row = seller_rows[-1]
    formula = "OR(" + ",".join(
        f'${seller_column}{start_row}="{seller}"' for seller in HIGHLIGHT_SELLERS
    ) + ")"
    formula_xml = escape(formula)
    rule_xml = (
        f'<conditionalFormatting sqref="{seller_column}{start_row}:{seller_column}{end_row}">'
        f'<cfRule type="expression" dxfId="{dxf_id}" priority="{priority}">'
        f"<formula>{formula_xml}</formula>"
        "</cfRule></conditionalFormatting>"
    )

    insertion_markers = [
        "<dataValidations",
        "<hyperlinks",
        "<printOptions",
        "<pageMargins",
        "<pageSetup",
        "<headerFooter",
        "<drawing",
        "<legacyDrawing",
        "<extLst",
        "</worksheet>",
    ]
    positions = [xml_text.find(marker) for marker in insertion_markers]
    positions = [position for position in positions if position >= 0]
    if not positions:
        raise ValueError("Koşullu biçimlendirme için çalışma sayfası ekleme noktası bulunamadı.")
    insert_at = min(positions)
    return xml_text[:insert_at] + rule_xml + xml_text[insert_at:]


def save_target_updates(
    source_path: Path,
    output_path: Path,
    sheet_name: str,
    updates: Dict[str, Any],
    seller_column: str,
) -> None:
    part_name = sheet_part_name(source_path, sheet_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}-",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(file_descriptor)

    try:
        with zipfile.ZipFile(source_path, "r") as source_zip, zipfile.ZipFile(temp_name, "w") as target_zip:
            styles_xml = source_zip.read("xl/styles.xml").decode("utf-8")
            styles_xml, yellow_dxf_id = ensure_yellow_dxf(styles_xml)
            for info in source_zip.infolist():
                data = source_zip.read(info.filename)
                if info.filename == "xl/styles.xml":
                    data = styles_xml.encode("utf-8")
                if info.filename == part_name:
                    xml_text = data.decode("utf-8")
                    xml_text = patch_sheet_xml(xml_text, updates)
                    xml_text = patch_seller_highlight_rule(
                        xml_text,
                        updates,
                        seller_column,
                        yellow_dxf_id,
                    )
                    data = xml_text.encode("utf-8")
                target_zip.writestr(info, data)
        os.replace(temp_name, output_path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def assert_only_target_columns(updates: Dict[str, Any], columns: Dict[str, int]) -> None:
    """Atanan Satıcı ve diğer statik sütunlara yazılmasını engeller."""
    allowed_columns = {
        get_column_letter(columns[SELLER_HEADER]),
        get_column_letter(columns[PRICE_HEADER]),
    }
    invalid_references = [
        reference
        for reference in updates
        if re.match(r"[A-Z]+", reference).group(0) not in allowed_columns
    ]
    if invalid_references:
        raise ValueError(
            "Statik bir sütuna yazma girişimi engellendi: "
            + ", ".join(invalid_references[:10])
        )


def open_scrape_browser(playwright, headful: bool):
    # Yerelde kurulu Google Chrome'u, Docker'da Playwright Chromium'u kullanır.
    # macOS'ta Chrome, Linux/Railway ortamında Playwright Chromium seçilir.
    default_browser_channel = "chrome" if sys.platform == "darwin" else ""
    browser_channel = os.getenv(
        "PLAYWRIGHT_BROWSER_CHANNEL",
        default_browser_channel,
    ).strip()
    launch_options = {
        "headless": not headful,
        "slow_mo": SLOW_MO,
    }
    if browser_channel:
        launch_options["channel"] = browser_channel
    browser = playwright.chromium.launch(**launch_options)
    context = browser.new_context(
        viewport={"width": 1440, "height": 1200},
        locale="tr-TR",
        service_workers="block",
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
    )

    # Ürün/satıcı bilgileri HTML ve sayfa JSON'undan okunur. Büyük görseller,
    # videolar ve fontlar veri doğruluğunu etkilemeden ağ ve render yükünü azaltır.
    def handle_route(route) -> None:
        if route.request.resource_type in {"image", "media", "font"}:
            route.abort()
        else:
            route.continue_()

    context.route("**/*", handle_route)
    page = context.new_page()
    page.set_default_timeout(PAGE_TIMEOUT_MS)
    return browser, page


def scrape_one_outcome(page, product_name: str) -> ScrapeOutcome:
    try:
        rows, summary = scrape_product(page, product_name)
        if rows and math.isfinite(rows[0].price_value):
            return ScrapeOutcome(
                product_name=product_name,
                seller_name=rows[0].seller_name,
                price_text=rows[0].price_text,
                price_value=rows[0].price_value,
                summary=summary,
                status="found",
            )
        return ScrapeOutcome(
            product_name=product_name,
            seller_name=None,
            price_text=None,
            price_value=None,
            summary=summary,
            status=(
                "verified_not_found"
                if summary.verified_no_match
                else "not_found"
            ),
        )
    except PlaywrightTimeoutError as exc:
        return ScrapeOutcome(
            product_name=product_name,
            seller_name=None,
            price_text=None,
            price_value=None,
            summary=ProductSummary(product_name, None, None, 0, None),
            status="timeout",
            error_message=str(exc),
        )
    except Exception as exc:
        return ScrapeOutcome(
            product_name=product_name,
            seller_name=None,
            price_text=None,
            price_value=None,
            summary=ProductSummary(product_name, None, None, 0, None),
            status="error",
            error_message=str(exc),
        )


def scrape_product_batch(products: List[str], headful: bool) -> List[ScrapeOutcome]:
    """Bir işçi sürecinde tek Chrome oturumuyla bir ürün grubunu tarar."""
    outcomes: List[ScrapeOutcome] = []
    with sync_playwright() as playwright:
        browser, page = open_scrape_browser(playwright, headful)
        try:
            for product_name in products:
                print(f"  → {product_name} aranıyor...", flush=True)
                outcome = scrape_one_outcome(page, product_name)
                outcomes.append(outcome)
                print(f"  ✓ {product_name} taraması tamamlandı.", flush=True)
        finally:
            browser.close()
    return outcomes


def split_product_batches(products: List[str], batch_count: int) -> List[List[str]]:
    """Uzun ve kısa ürünleri işçilere dengeli dağıtmak için sıralı dönüşümlü böler."""
    batches = [[] for _ in range(batch_count)]
    for index, product_name in enumerate(products):
        batches[index % batch_count].append(product_name)
    return [batch for batch in batches if batch]


def write_result(
    updates: Dict[str, Any],
    columns: Dict[str, int],
    row_numbers: List[int],
    seller: str,
    price: Optional[float],
) -> None:
    seller_column = get_column_letter(columns[SELLER_HEADER])
    price_column = get_column_letter(columns[PRICE_HEADER])
    for row_number in row_numbers:
        updates[f"{seller_column}{row_number}"] = seller
        updates[f"{price_column}{row_number}"] = price


def apply_scrape_outcome(
    outcome: ScrapeOutcome,
    completed_count: int,
    total_count: int,
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    columns: Dict[str, int],
    model_rows: Dict[str, List[int]],
    updates: Dict[str, Any],
) -> Tuple[int, int, int]:
    product = outcome.product_name
    prefix = f"[{completed_count}/{total_count}] {product} ... "

    if outcome.status == "found":
        write_result(
            updates,
            columns,
            model_rows[product],
            outcome.seller_name or "Bulunamadı",
            outcome.price_value,
        )
        print(
            prefix
            + f"{outcome.seller_name} / {outcome.price_text} "
            + f"| seçilen ilan: {outcome.summary.rating_count} değerlendirme "
            + f"| {outcome.summary.seller_count} satıcı karşılaştırıldı"
        )
        counts = (1, 0, 0)
    elif outcome.status in ("not_found", "verified_not_found"):
        write_result(updates, columns, model_rows[product], "Bulunamadı", None)
        print(prefix + "bulunamadı")
        counts = (0, 1, 0)
    elif outcome.status == "timeout":
        write_result(updates, columns, model_rows[product], "Zaman aşımı", None)
        print(prefix + "zaman aşımı")
        counts = (0, 0, 1)
    else:
        write_result(updates, columns, model_rows[product], "Hata", None)
        print(prefix + f"hata: {outcome.error_message or 'bilinmeyen hata'}")
        counts = (0, 0, 1)

    # Uzun listelerde kesinti olsa bile tamamlanan sonuçlar kaybolmasın.
    assert_only_target_columns(updates, columns)
    save_target_updates(
        input_path,
        output_path,
        sheet_name,
        updates,
        get_column_letter(columns[SELLER_HEADER]),
    )
    return counts


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Kod içindeki Model listesini Hepsiburada'da arar ve yalnızca "
            "'En iyi HB Satıcı' ile 'En iyi HB Satıcı Fiyat' sütunlarını günceller."
        )
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help=(
            "Oluşturulacak .xlsx dosyası. Belirtilmezse kodun başındaki "
            "OUTPUT_DIRECTORY ve OUTPUT_FILENAME kullanılır."
        ),
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Ürün listesinin bulunduğu sayfa")
    parser.add_argument("--headful", action="store_true", help="Chrome penceresini görünür aç")
    parser.add_argument(
        "--model",
        help="Yalnızca belirtilen modeli işle (örnek: A159W-N1DF)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Yalnızca ilk N benzersiz modeli işle (test için)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_WORKERS,
        help=(
            f"Aynı anda çalışacak tarayıcı sayısı (1-{MAX_WORKERS}, "
            f"varsayılan: {DEFAULT_WORKERS})"
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not 1 <= args.workers <= MAX_WORKERS:
        raise ValueError(f"--workers 1 ile {MAX_WORKERS} arasında olmalıdır.")
    input_path = create_embedded_template_file()
    atexit.register(lambda: input_path.unlink(missing_ok=True))
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else configured_output_path()
    )

    workbook, sheet, columns, model_rows, updates = prepare_workbook(
        input_path=input_path,
        output_path=output_path,
        sheet_name=args.sheet,
    )
    products = list(model_rows)
    if args.model:
        target_model = normalize_model_code(args.model)
        products = [p for p in products if normalize_model_code(p) == target_model]
        if not products:
            raise ValueError(f"Excel'de model bulunamadı: {args.model}")
    elif args.limit is not None:
        if args.limit < 1:
            raise ValueError("--limit en az 1 olmalıdır.")
        products = products[:args.limit]

    print(f"\n{len(products)} benzersiz model Hepsiburada'da aranacak.")
    print("Statik ürün listesi kod içindeki Excel şablonundan oluşturuldu.")
    print(f"Çıktı dosyası: {output_path}")

    worker_count = min(args.workers, len(products))
    if args.headful and worker_count > 1:
        worker_count = 1
        print("Görünür Chrome modunda karışıklığı önlemek için 1 işçi kullanılacak.")
    print(f"Paralel tarayıcı sayısı: {worker_count}\n")

    found_count = 0
    not_found_count = 0
    error_count = 0
    completed_count = 0

    def record_outcome(outcome: ScrapeOutcome) -> None:
        nonlocal completed_count, found_count, not_found_count, error_count
        completed_count += 1
        found, not_found, error = apply_scrape_outcome(
            outcome=outcome,
            completed_count=completed_count,
            total_count=len(products),
            input_path=input_path,
            output_path=output_path,
            sheet_name=args.sheet,
            columns=columns,
            model_rows=model_rows,
            updates=updates,
        )
        found_count += found
        not_found_count += not_found
        error_count += error

    if worker_count == 1:
        with sync_playwright() as playwright:
            browser, page = open_scrape_browser(playwright, args.headful)
            try:
                for product in products:
                    print(f"  → {product} aranıyor...", flush=True)
                    record_outcome(scrape_one_outcome(page, product))
            finally:
                browser.close()
    else:
        # Her işçi tek Chrome oturumunu kendi ürün grubunda yeniden kullanır.
        # Ek küçük paketler yeni tarayıcı açma/kapatma maliyetini artırıyordu.
        batch_count = min(len(products), worker_count)
        batches = split_product_batches(products, batch_count)
        retry_products: List[str] = []

        with ProcessPoolExecutor(max_workers=worker_count) as executor:
            future_batches = {
                executor.submit(scrape_product_batch, batch, False): batch
                for batch in batches
            }
            for future in as_completed(future_batches):
                batch = future_batches[future]
                try:
                    for outcome in future.result():
                        if outcome.status in ("found", "verified_not_found"):
                            record_outcome(outcome)
                        else:
                            # Paralel oturumda geçici engel/zaman aşımı olasılığına
                            # karşı sonuçsuz ürünleri tek tarayıcıda tekrar doğrula.
                            retry_products.append(outcome.product_name)
                except Exception as exc:
                    print(
                        f"Bir paralel tarayıcı kapanmış ({exc}); "
                        f"{len(batch)} ürün güvenli biçimde yeniden denenecek."
                    )
                    retry_products.extend(batch)

        # Bir işçi süreci çökerse doğruluktan ödün vermeden ürünleri tek oturumda
        # tekrar tara; sonuçsuz bırakma.
        retry_products = list(dict.fromkeys(retry_products))
        if retry_products:
            print(f"{len(retry_products)} sonuçsuz ürün tek oturumda yeniden doğrulanacak.")
            with sync_playwright() as playwright:
                browser, page = open_scrape_browser(playwright, False)
                try:
                    for product in retry_products:
                        record_outcome(scrape_one_outcome(page, product))
                finally:
                    browser.close()

    workbook.close()
    input_path.unlink(missing_ok=True)

    print("\nTamamlandı.")
    print(f"Bulunan: {found_count} | Bulunamayan: {not_found_count} | Hata: {error_count}")
    print(f"Dosya: {output_path}\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nİşlem kullanıcı tarafından durduruldu.")
        sys.exit(1)
